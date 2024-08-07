from PyQt5.QtWidgets import     QApplication, QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QHBoxLayout, QCheckBox, QGroupBox, QScrollArea, QScrollBar, QSplitter, QFrame
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import pyqtSlot, Qt
import sys
import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
import matplotlib
matplotlib.use('Qt5Agg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg, NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from matplotlib.widgets import MultiCursor, Cursor, TextBox

class MplCanvas(FigureCanvasQTAgg):
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = self.fig.add_subplot(111)
        super(MplCanvas, self).__init__(self.fig)
    def add_legend(self, legend_labels):
        self.axes.legend(legend_labels)

class Main(QWidget):
    def __init__(self):
        super(Main, self).__init__()
        self.setWindowTitle("Loading Excel")
        self.checkboxes=[]
        self.number_of_grafics = 0
        self.canvases = []
        self.legend = []
        self.my_blocks = []
        self.lay=[]
        self.textt = []
        self.number_of_cursors = 0
        layout = QVBoxLayout() 
        self.manygrafics = QSplitter(Qt.Vertical)
        self.boxes = []
        self.setLayout(layout)
        self.table_widget = QTableWidget()
        layout.addLayout(self.button_block())   #
        layout.addWidget(self.manygrafics)      #
        tablesplitter = QSplitter(Qt.Vertical)
        tablesplitter.addWidget(self.manygrafics)
        tablesplitter.addWidget(self.table_widget)
        layout.addWidget(tablesplitter)     #
        self.on_click3()

    def button_block(self):        #основные кнопки верхней панели
        block = QHBoxLayout()
        self.button1 = QPushButton("Открыть новый файл")
        self.button2 = QPushButton("Обновить график")
        self.button3 = QPushButton("Добавить координаты")
        self.button4 = QPushButton("Удалить координаты")
        self.button5 = QPushButton("Настройки курсора")
        self.button1.clicked.connect(self.on_click)
        self.button2.clicked.connect(self.on_click2)
        self.button3.clicked.connect(self.on_click3)
        self.button4.clicked.connect(self.delete_click)
        self.button5.clicked.connect(self.cursor_settings)
        block.addWidget(self.button1)
        block.addWidget(self.button2)
        block.addWidget(self.button3)
        block.addWidget(self.button4)
        self.button2.setEnabled(False)
        self.button4.setEnabled(False)
        #block.addWidget(self.button5)
        return block
        
    @pyqtSlot()
    def on_click(self):
        name, done1 = QtWidgets.QInputDialog.getText(
			self, 'Input Dialog', 'Введите название или путь к нужному файлу:') 
        if done1: 
            self.load_data(name)
            self.make_checkboxes()
            self.clear_canvases()
            
    def on_click2(self):
        for i in self.canvases: i.axes.clear()
        self.cursors = []
        self.Main_sheet = list(reversed(list(zip(*self.list_values[1:]))))
        Axes = []
        for graph in range(self.number_of_grafics):
            my_legend = []
            for i in range(len(self.legend) - 1):
                if self.checkboxes[graph][i].isChecked():
                    self.canvases[graph].axes.plot(self.Main_sheet[-1], self.Main_sheet[-1*(2+i)], label = self.legend[i+1])
                    my_legend.append(self.legend[i+1])
            self.canvases[graph].add_legend(my_legend)
            Axes.append(self.canvases[graph].axes)
            self.canvases[graph].draw()
            self.canvases[graph].installEventFilter(self)
        self.Axes = tuple(Axes)
        self.multi = MultiCursor(self.canvases[-1].fig, self.Axes, horizOn=False, vertOn=True, useblit=True, color='red') 
        self.canvases[-1].draw()
        
    def eventFilter(self, obj, event):
        if obj in self.canvases:
            if event.type() == event.MouseButtonPress:
                if event.button() == Qt.LeftButton:  # Левая кнопка мыши
                    self.number_of_cursors += 1
                    delta = int(self.Main_sheet[-1][-1] )
                    #print(delta)
                    my_x = 0
                    for i in self.Main_sheet[-1]:
                        if abs(int(i) - self.xpos) < delta:
                            delta = abs(int(i) - self.xpos)
                            my_x = i
                    num_of_x = list(self.Main_sheet[-1]).index(my_x)
                    #self.xpos
                    for i in range(len(self.canvases)):
                        textt = "x: " + str(my_x) 
                        for j in range(len(self.legend) - 1):
                            if self.checkboxes[i][j].isChecked():                   #Здесь должно находиться составление текста для надписи курсора
                                #print(num_of_x)
                                #print(self.legend)
                                #print(self.Main_sheet)
                                textt += "\n" + str(self.legend[j+1]) + ": " + str(list(self.Main_sheet[-1*(2+j)])[num_of_x])
                        #plt.text(0.5, 0.5, 'Center', transform = i.axes.transAxes)
                        ymi = self.canvases[i].axes.get_ylim()[0] + (self.canvases[i].axes.get_ylim()[1] - self.canvases[i].axes.get_ylim()[0]) / 20
                        yma = self.canvases[i].axes.get_ylim()[1] - (self.canvases[i].axes.get_ylim()[1] - self.canvases[i].axes.get_ylim()[0]) / 20
                        self.canvases[i].axes.plot([self.xpos, self.xpos], [ymi, yma], color = 'darkblue')
                        #i.axes.annotate('hi', xy=(self.xpos, self.ypos))
                        self.textt.append(self.canvases[i].axes.text(self.xpos, (self.canvases[i].axes.get_ylim()[0] + self.canvases[i].axes.get_ylim()[1]) /2, textt, va='center', ha='left'))
                        self.canvases[i].draw()
                    self.multi.visible = False  # Остановка курсора
                    ###
                elif event.button() == Qt.RightButton:  # Правая кнопка мыши
                    self.multi.visible = True  # Возврат курсора в исходное состояние
                    for i in self.textt:
                        i.visible = False
                    for i in self.canvases: i.draw()
                    
        return super().eventFilter(obj, event)        

    def on_click3(self):
        self.number_of_grafics += 1
        self.my_blocks.append(QGroupBox())
        grafic = QSplitter(Qt.Horizontal)
        self.lay.append(QHBoxLayout())
        self.my_blocks[-1].setLayout(self.lay[-1])
        self.lay[-1].addWidget(grafic)
        canv1 = QGroupBox()
        canv = QVBoxLayout()
        canv1.setLayout(canv)
        checkboxes1 = QGroupBox()
        checkboxes1.setTitle("График №"+str(self.number_of_grafics))
        self.boxes.append(QVBoxLayout())
        self.checkboxes.append([])
        for i in range(len(self.legend) - 1):
                c = QCheckBox(self.legend[i+1], self)
                c.clicked.connect(self.on_click2)
                self.checkboxes[-1].append(c)
                self.boxes[-1].addWidget(self.checkboxes[-1][i])
        checkboxes1.setLayout(self.boxes[-1])
        self.canvases.append(MplCanvas(self, width=5, height=4, dpi=100))
        #self.canvases[-1].mpl_connect("button_press_event", self.on_press)
        #self.canvas.mpl_connect("button_release_event", self.on_release)
        scroll = QScrollArea()
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setWidgetResizable(True)
        scroll.setWidget(checkboxes1)
        self.canvases[-1].mpl_connect("motion_notify_event", self.on_move)
        toolbar = NavigationToolbar(self.canvases[self.number_of_grafics-1], self)
        canv.addWidget(toolbar)
        canv.addWidget(self.canvases[self.number_of_grafics-1])
        grafic.addWidget(canv1)
        grafic.addWidget(scroll)
        self.manygrafics.addWidget(self.my_blocks[-1])
        self.button4.setEnabled(True)

    def on_move(self, event):
        self.xpos = event.xdata
        self.ypos = event.ydata
        if self.xpos != None and self.ypos != None: self.my_text = 'x: ' + str(round(self.xpos, 3)) + '\n y:' + str(round(self.ypos, 3))
        else: self.my_text = ' '
        #t.delete()
        #t = self.canvases[-1].axes.text(self.xpos, self.ypos, self.text,
        #    ha="left", va="top", rotation=0, size=15,
        #    bbox=dict(boxstyle="square,pad=0.3",
        #              fc="lightblue", ec="steelblue", lw=2))
        #t.set_visible(False)
        #self.canvases[-1].draw()    
        
    def delete_click(self):
        layout = self.lay[-1]
        for i in reversed(range(layout.count())):
            widgetToRemove = layout.itemAt(i).widget()
            layout.removeWidget(widgetToRemove)
            widgetToRemove.setParent(None)
        self.my_blocks[-1].setParent(None)
        self.canvases = self.canvases[:-1]
        self.checkboxes = self.checkboxes[:-1]
        self.boxes = self.boxes[:-1]
        self.lay = self.lay[:-1]
        self.my_blocks = self.my_blocks[:-1]
        self.Axes = self.Axes[:-1]
        self.multi = MultiCursor(self.canvases[-1].fig, self.Axes, horizOn=False, vertOn=True, useblit=True, color='red')
        self.number_of_grafics -=1
        for i in self.canvases: i.draw()
        if self.number_of_grafics == 1: self.button4.setEnabled(False)
        
    def cursor_settings(self):
        self.multi = MultiCursor(self.canvases[-1].fig, self.Axes, horizOn=False, vertOn=True, useblit=True, color='green')
        for i in self.canvases: i.draw()
        #print(self.canvases[-1].axes.get_xlim()[1], self.canvases[-1].axes.get_ylim()[1])
        

    def clear_canvases(self):
        for i in range(self.number_of_grafics):
            self.canvases[i].axes.clear()

    def load_data(self, path):
        workbook = None
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        self.table_widget.setRowCount(sheet.max_row - 1)
        self.table_widget.setColumnCount(sheet.max_column)

        self.list_values = list(sheet.values)
        self.legend = list(self.list_values[0])                                   #self.legeng - матрица со всеми названиями, включая ось х
        self.table_widget.setHorizontalHeaderLabels(self.legend)

        row_index = 0
        for value_tuple in self.list_values[1:]:
            column_index = 0
            for value in value_tuple:
                self.table_widget.setItem(row_index, column_index,QTableWidgetItem(str(value)))
                column_index += 1
            row_index += 1
        self.button2.setEnabled(True)

    def make_checkboxes(self):
        for graph in range(self.number_of_grafics):
            for i in self.checkboxes[graph]: self.boxes[graph].removeWidget(i)                
        self.checkboxes =[]
        for graph in range(self.number_of_grafics):
            self.checkboxes.append([])
            for i in range(len(self.legend) - 1):
                c = QCheckBox(self.legend[i+1], self)
                c.clicked.connect(self.on_click2)
                self.checkboxes[graph].append(c)
                self.boxes[graph].addWidget(self.checkboxes[graph][i])


app = QApplication(sys.argv)
window = Main()
window.showMaximized()
app.exec_()